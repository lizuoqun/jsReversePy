'''
实时票房网页
https://piaofang.maoyan.com/box-office?ver=normal
'''

import requests
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

cookies = {
    '_lxsdk_cuid': '194de2c5d44c8-082648cccc4b87-26011b51-1bcab9-194de2c5d44c8',
    '_lxsdk': '194de2c5d44c8-082648cccc4b87-26011b51-1bcab9-194de2c5d44c8',
    'uuid': '194de2c5d44c8-082648cccc4b87-26011b51-1bcab9-194de2c5d44c8',
    '_lx_utm': 'utm_source%3DBaidu%26utm_medium%3Dorganic',
    '_lxsdk_s': '194de2c5d44-fc-dfe-260%7C%7C22',
}

headers = {
    'Accept': 'application/json, text/plain, */*',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'keep-alive',
    # 'Cookie': '_lxsdk_cuid=194de2c5d44c8-082648cccc4b87-26011b51-1bcab9-194de2c5d44c8; _lxsdk=194de2c5d44c8-082648cccc4b87-26011b51-1bcab9-194de2c5d44c8; uuid=194de2c5d44c8-082648cccc4b87-26011b51-1bcab9-194de2c5d44c8; _lx_utm=utm_source%3DBaidu%26utm_medium%3Dorganic; _lxsdk_s=194de2c5d44-fc-dfe-260%7C%7C22',
    'Referer': 'https://piaofang.maoyan.com/box-office?ver=normal',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'same-origin',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0.0.0 Safari/537.36',
    'sec-ch-ua': '"Not A(Brand";v="8", "Chromium";v="132", "Google Chrome";v="132"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
    'uid': '4551f77ebc2b218e14e89d83f2c49b4b6b2600af',
}

params = {
    'date': '1',
    'isSplit': 'true',
    'allowAd': '1',
    'timeStamp': '1738894308412',
    'User-Agent': 'TW96aWxsYS81LjAgKFdpbmRvd3MgTlQgMTAuMDsgV2luNjQ7IHg2NCkgQXBwbGVXZWJLaXQvNTM3LjM2IChLSFRNTCwgbGlrZSBHZWNrbykgQ2hyb21lLzEzMi4wLjAuMCBTYWZhcmkvNTM3LjM2',
    'index': '447',
    'channelId': '40009',
    'sVersion': '2',
    'signKey': 'a3cf5985a740ad8c189866a9c58d54ce',
    'WuKongReady': 'h5',
}

response = requests.get('https://piaofang.maoyan.com/i/api/getBoxList', params=params, cookies=cookies,
                        headers=headers)

data = response.json()['boxOffice']['data']['list']

# 将嵌套的字典展开为平面结构
flattened_data = []
for item in data:
    movie_info = item['movieInfo']
    flattened_item = {
        '电影id': movie_info['movieId'],
        '电影名称': movie_info['movieName'],
        '综合票房（万元）': item['boxDesc'],
        '综合票房占比': item['boxRate'],
        '上映时间': movie_info['releaseInfo'],
        '排座占比': item['seatCountRate'],
        '排片占比': item['showCountRate'],
        '总票房': item['sumBoxDesc']
    }
    print('----', flattened_item)
    flattened_data.append(flattened_item)

# 创建DataFrame
df = pd.DataFrame(flattened_data)

excel_file = 'movieRank.xlsx'

# 保存到Excel文件
df.to_excel(excel_file, index=False, engine='openpyxl')

# 使用openpyxl加载工作簿并调整列宽
wb = load_workbook(excel_file)
ws = wb.active

# 设置列宽（可以根据需要调整）
column_widths = {
    'A': 10,
    'B': 30,
    'C': 20,
    'D': 15,
    'E': 15,
    'F': 15,
    'G': 15,
    'H': 15
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# 设置单元格居中对齐
alignment = Alignment(horizontal='center', vertical='center')
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    for cell in row:
        cell.alignment = alignment

# 设置表头背景色为绿色
header_fill = PatternFill(start_color="0eb0c9", end_color="00FF00", fill_type="solid")
for cell in ws[1]:
    cell.fill = header_fill

# 保存修改后的工作簿
wb.save(excel_file)

print("数据已成功写入Excel文件")
